import pip, os, sys, io, re, time
import random
import json

from threading import Thread
from pip._internal import main as pipmain

sys.path.append("/usr/local/lib/python3.6/dist-packages")

def install(package):
    pipmain(['install', '--no-cache-dir', package])

def uninstall(package):
    pipmain(['uninstall', package])

try:
    import pandas as pd
    import geocoder
    import openpyxl
    import requests
    import time
    import math
    import numpy as np
    import IPython.display as IPython
    install('timeit')
    import timeit
    print('package imports are good')
except ImportError:
    #install('pip')
    install('pandas')
    import pandas as pd
    try:
        !pip install geocoder
        print('Open a terminal and install geocoder using pip')
    except:
        print('Open a terminal and install geocoder using pip')
        sys.exit()
    import geocoder
    install('openpyxl')
    import openpyxl
    install('numpy')
    import numpy as np
    import math
    import requests
    import time
    install('IPython.display')
    import IPython.display as IPython
    install('timeit')
    import timeit
    print('completed package installs. IF THERE ARE ANY ERRORS open a terminal session, upgrade pip and install the package needed, and try again')

from IPython.display import clear_output
from pandas import ExcelWriter, ExcelFile
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

poss_regex = r"(?<=[a-z])[\']([A-Z])"
total = 0
start = None
threads = []
mapbox_key = "pk.eyJ1Ijoiam93ZWluZXIiLCJhIjoiY2p5ZnZkaG9kMGFlMTNkcGExZTdicjlkcyJ9.JrOaDB18w4UOAOrO1lOY6A"
here_id = "qn47ynVwJoIj3kGf4SDk"
here_code = "aYhCOxaGuROnuY0KNkUrRA"

def capitalizeWords(title):
    global poss_regex
    if title is not None:
        # Take advantage of title(), we'll fix the apostrophe issue afterwards
        # Special handling for contractions
#         return re.sub(poss_regex, lowercase_match_group, title.title())
        return title.upper()

def getCoordsArcGis(location):
    g = geocoder.arcgis(location)
    coord_json = g.json
    lat = coord_json['lat']
    long = coord_json['lng']
    #print(coord_json)
    #print(str(test) + ":  " + str(lat) + ", " + str(long))
    l = []
    l.append(lat)
    l.append(long)
    return l

def async_request_three(location, ws, session, index, locations_list, loc_col_index, lat_col_index, long_col_index):
    global total, threads, here_id, here_code, total_threads
    time.sleep((random.random() * 10))
    try:
        arc = geocoder.here(location, app_id = here_id, app_code = here_code)
        l = []
        if arc.ok:
#             l = arc.latlng
    #         l.append(arc.json[0])
    #         l.append(arc.json[1])
            l.append(arc.json['lat'])
            l.append(arc.json['lng'])
            cell = ws.cell(row=index+2, column=loc_col_index + 1)
            name = str(locations_list[index])
            if name is not None and name != '' and name != '' and not pd.isnull(name):
                cell.value = capitalizeWords(str(locations_list[index]))
            else:
                cell.value = ''
            cell = ws.cell(row=index + 2, column=lat_col_index + 1)
            cell.value = l[0]
            cell = ws.cell(row=index + 2, column=long_col_index + 1)
            cell.value = l[1]
        else:
            Thread(target=async_request, args=(location, ws, session, index, locations_list, loc_col_index, lat_col_index, long_col_index))
            threads.append(t)
            t.start()
            total += 1
            total_threads += 1
    except:
        t = Thread(target=async_request, args=(location, ws, session, index, locations_list, loc_col_index, lat_col_index, long_col_index))
        threads.append(t)
        t.start()
        total += 1
        total_threads += 1

def async_request_two(location, ws, session, index, locations_list, loc_col_index, lat_col_index, long_col_index):
    global total, threads, total_threads
    time.sleep(random.random() * 10)
    try:
        arc = geocoder.arcgis(location)
        l = []
        if arc.ok:
#             l = arc.latlng
    #         l.append(arc.json[0])
    #         l.append(arc.json[1])
            l.append(arc.json['lat'])
            l.append(arc.json['lng'])
            cell = ws.cell(row=index+2, column=loc_col_index + 1)
            name = str(locations_list[index])
            if name is not None and name != '' and name != '' and not pd.isnull(name):
                cell.value = capitalizeWords(str(locations_list[index]))
            else:
                cell.value = ''
            cell = ws.cell(row=index + 2, column=lat_col_index + 1)
            cell.value = l[0]
            cell = ws.cell(row=index + 2, column=long_col_index + 1)
            cell.value = l[1]
    #         clear_output(wait=True)
        else:
            Thread(target=async_request_three, args=(location, ws, session, index, locations_list, loc_col_index, lat_col_index, long_col_index))
            threads.append(t)
            t.start()
            total += 1
            total_threads += 1
    except:
        t = Thread(target=async_request_three, args=(location, ws, session, index, locations_list, loc_col_index, lat_col_index, long_col_index))
        threads.append(t)
        t.start()
        total += 1
        total_threads += 1

def async_request(location, ws, session, index, locations_list, loc_col_index, lat_col_index, long_col_index):
    global total, threads, mapbox_key, total_threads
    time.sleep((random.random() * 10))
    try:
        arc = geocoder.mapbox(location, key=mapbox_key)
        l = []
#         l = arc["features"]["center"]
        if arc.ok:
#         l.append(arc.json[0])
#         l.append(arc.json[1])
            l.append(arc.json["lat"])
            l.append(arc.json["lng"])
            cell = ws.cell(row=index+2, column=loc_col_index + 1)
            name = str(locations_list[index])
            if name is not None and name != '' and name != '' and not pd.isnull(name):
                cell.value = capitalizeWords(str(locations_list[index]))
            else:
                cell.value = ''
            cell = ws.cell(row=index + 2, column=lat_col_index + 1)
            cell.value = l[0]
            cell = ws.cell(row=index + 2, column=long_col_index + 1)
            cell.value = l[1]
        else:
            t = Thread(target=async_request_two, args=(location, ws, session, index, locations_list, loc_col_index, lat_col_index, long_col_index))
            threads.append(t)
            total_threads += 1
            t.start()
            total += 1
    except:
        t = Thread(target=async_request_two, args=(location, ws, session, index, locations_list, loc_col_index, lat_col_index, long_col_index))
        threads.append(t)
        t.start()
        total += 1
        total_threads += 1
        #     except:
#         time.sleep(random.random())
#         Thread(target=async_request, args=(location, ws, session, index, locations_list, loc_col_index, lat_col_index, long_col_index)).start()

total_threads = 0

def test_req(location, key, here_id, here_code, n):
#     if n % 2 == 0:
#         arc = geocoder.mapbox(location, key=key)
#         l = []
#         if arc.ok:
#             l.append(arc.json["lat"])
#             l.append(arc.json["lng"])
#     else:
    arc = geocoder.osm(location)
    l = []
    if arc.ok:
        l.append(arc.json['lat'])
        l.append(arc.json['lng'])
    return

#maintain open session to increase rate of requests --> fill file faster
def session_requests(addresses, locations, df, wb, ws, file):
    global total, total_threads, mapbox_key, here_id, here_code
    total = 0
    locations = locations
    loc_col_index = df.columns.get_loc('Location')
    lat_col_index = df.columns.get_loc("Latitude")
    long_col_index = df.columns.get_loc("Longitude")
    zip_col_index = df.columns.get_loc("Zipcode")
    start = timeit.default_timer()
    with requests.Session() as session:
        i = 0
        for int, val in enumerate(addresses):
            if isinstance(val, str) and val is not None and not pd.isnull(str(val)) and val.find('-State') < 0:
                print(val)
                clear_output(wait=True)
#                 if i % 3 == 0:
#                     t = Thread(target=async_request_three, args=(val, ws, session, i, locations, loc_col_index, lat_col_index, long_col_index))
#                 elif i % 4 == 0:
#                     t = Thread(target=async_request_two, args=(val, ws, session, i, locations, loc_col_index, lat_col_index, long_col_index))
#                 else:
#                     t = Thread(target=async_request, args=(val, ws, session, i, locations, loc_col_index, lat_col_index, long_col_index))
                time.sleep(0.03)
                t = Thread(target=async_request, args=(val, ws, session, i, locations, loc_col_index, lat_col_index, long_col_index))
                threads.append(t)
                t.start()
                total_threads += 1
                total += 1
                #Zipcode processing
                cell = ws.cell(row=i+2, column=zip_col_index + 1)
                address = str(val)
                address_list = address.split(' ')
                zip_code = address_list[-1]
                if zip_code[0].isdigit():
                    cell.value = zip_code
            prog = (i / len(addresses)) * 100
            stop = timeit.default_timer()
            clear_output(wait=True)
            print('Creating all threads [%d%%]\r'%(prog))
            print('Current runtime:', np.round((stop-start),2), "seconds")
            i += 1
    for x in threads:
        clear_output(wait=True)
        stop = timeit.default_timer()
        print('Current runtime:', np.round((stop-start),2), "seconds")
        print('Threads remaining:', total)
        x.join()
        clear_output(wait=True)
        total -= 1
    wb.save(file)
    stop = timeit.default_timer()
    clear_output(wait=True)
    new_start = timeit.default_timer()
    n = 0
    while n < 10:
        print("Comparing with non-optimized requests")
        print(n + 1, "/", 10)
        test_req("1 University Square Drive, Princeton, NJ 08540", mapbox_key, here_id, here_code, n)
        n += 1
        clear_output(wait=True)
    new_stop = timeit.default_timer()
    new_t = (new_stop - new_start) / 10
    clear_output(wait=True)
    print("Finished!")
    print("All",total_threads, "threads exexuted with time:", np.round((stop-start) / 60,0),"minutes", np.round((stop-start) % 60 ,2),"seconds")
    print("Predicted non-optimized expected execution time:", np.round((new_t*total_threads / 60), 0),"minutes",np.round((new_t*total_threads % 60), 2), "seconds")
    print("Avereage time per thread:", np.round(((stop-start) / total_threads), 4), "seconds")
    print("Non-optimized time per thread:", np.round(new_t, 4), "seconds")

def create_list(df, wb, ws, file):
    global total, mapbox_key, here_id, here_code
    i = 0
    addresses = []
    locations = []
#     start = time.time()
    while(i < len(df['Address'])):
        addresses.append(df['Address'][i])
        locations.append(df['Location'][i])
        i += 1
#     end = time.time()
    session_requests(addresses, locations, df, wb, ws, file)
#     print(addresses)
#     print(end - start)

def open_file():
    global total, mapbox_key, here_id, here_code
    directory = '.'
    os.chdir(directory)
    df = None
    del df
    #print(os.listdir())
    #Loop and retrieve file to assign to dataframe
    for file in os.listdir():
        if file.find('multi state') >= 0 and file.find('.xlsx') >= 0:
            #f = wlrd
            print(file)
            workbook = load_workbook(file)
            worksheets = workbook.sheetnames #return a list of sheetnames in the workbook
            #print(worksheets[1])
            #Access the second worksheet to read from and write to
            worksheet = workbook[worksheets[1]]
            df = pd.read_excel(file, sheet_name=1)
            create_list(df, workbook, worksheet, file)

open_file()
